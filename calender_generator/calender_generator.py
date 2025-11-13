import json
import calendar
import datetime
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuration ---

CONFIG_FILE = 'config.json'
OUTPUT_FILE = 'Mosaic_AI_Release_Calendar.xlsx'
DAY_ABBREVIATIONS = ["S", "M", "T", "W", "T", "F", "S"]

# Map color names from JSON to Excel-compatible HEX color codes (AARRGGBB format)
# Using light, print-friendly colors.
COLOR_MAP = {
    "blue": "FFADD8E6",    # Light Blue
    "green": "FFC6E0B4",   # Light Green
    "red": "FFFFC7CE",     # Light Red (Pinkish)
    "yellow": "FFFFFF99",  # Light Yellow
    "purple": "FFD8BFD8",  # Light Purple (Thistle)
    "gray": "FFF0F0F0",    # Light Gray (for non-month days)
}

# --- Styling Definitions ---

def get_styles():
    """Returns a dictionary of reusable openpyxl styles."""
    
    # --- Fonts ---
    main_title_font = Font(size=22, bold=True, name="Calibri", color="FFFFFFFF") 
    month_title_font = Font(size=14, bold=True, name="Calibri", color="FFFFFFFF")
    legend_title_font = Font(size=14, bold=True, name="Calibri")
    day_header_font = Font(size=10, bold=True, name="Calibri")
    date_font = Font(size=10, name="Calibri")
    weekend_font = Font(size=10, name="Calibri", color="FF808080") # Gray text for weekends
    
    # --- Alignment ---
    center_align = Alignment(horizontal='center', vertical='center')
    top_right_align = Alignment(horizontal='right', vertical='top', wrap_text=True)

    # --- Borders ---
    thin_side = Side(style='thin', color='FFD0D0D0') # Light gray border
    date_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # --- Fills ---
    color_fills = {name: PatternFill(start_color=hex_val, end_color=hex_val, fill_type="solid")
                   for name, hex_val in COLOR_MAP.items()}
    heading_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")

    return {
        "main_title_font": main_title_font,
        "month_title_font": month_title_font,
        "legend_title_font": legend_title_font,
        "day_header_font": day_header_font,
        "date_font": date_font,
        "weekend_font": weekend_font,
        "center_align": center_align,
        "top_right_align": top_right_align,
        "date_border": date_border,
        "fills": color_fills,
        "heading_fill": heading_fill,
    }

# --- Core Logic ---

def load_config(filename):
    """Loads and validates the main config file."""
    if not os.path.exists(filename):
        print(f"Error: Configuration file '{filename}' not found.")
        sys.exit(1)
    
    try:
        with open(filename, 'r') as f:
            config = json.load(f)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON in '{filename}'. {e}")
        sys.exit(1)
        
    if 'calendar_year' not in config:
        print("Error: 'calendar_year' not found in config file.")
        sys.exit(1)
        
    return config

def process_events(config):
    """
    Processes the config into an easy-to-use date-to-color map
    and a legend map.
    Handles single dates (dd/mm/yyyy), short dates (dd/mm),
    and date ranges (dd/mm-dd/mm).
    
    *** Date ranges ignore Saturdays and Sundays. ***
    """
    event_map = {}
    legend_data = {}
    year = config['calendar_year']

    for key, value in config.items():
        if key == 'calendar_year':
            continue
            
        if isinstance(value, dict) and 'dates' in value and 'colour_code' in value:
            occasion_name = key.replace('_', ' ').title()
            color = value['colour_code']
            
            if color not in COLOR_MAP:
                print(f"Warning: Color '{color}' for '{occasion_name}' not recognized. Skipping.")
                continue
                
            if occasion_name not in legend_data:
                legend_data[occasion_name] = color
            
            for date_str in value['dates']:
                try:
                    if '-' in date_str:
                        # Case 1: Date Range ("dd/mm-dd/mm")
                        start_str, end_str = date_str.split('-')
                        
                        start_date = datetime.datetime.strptime(f"{start_str}/{year}", '%d/%m/%Y').date()
                        end_date = datetime.datetime.strptime(f"{end_str}/{year}", '%d/%m/%Y').date()

                        if start_date > end_date:
                            print(f"Warning: Invalid date range '{date_str}' (start is after end). Skipping.")
                            continue

                        # Loop from start to end (inclusive)
                        current_date = start_date
                        while current_date <= end_date:
                            # MODIFIED: Check if the day is not Saturday (5) or Sunday (6)
                            # 0=Monday, 1=Tuesday, ..., 5=Saturday, 6=Sunday
                            if current_date.weekday() not in [5, 6]:
                                # Use existing color if a more specific event (e.g., holiday) is already set
                                if current_date.strftime('%d/%m/%Y') not in event_map:
                                    event_map[current_date.strftime('%d/%m/%Y')] = color
                            
                            # Always increment the day
                            current_date += datetime.timedelta(days=1)
                            
                    elif len(date_str.split('/')) == 3:
                        # Case 2: Full Single Date ("dd/mm/yyyy")
                        event_date = datetime.datetime.strptime(date_str, '%d/%m/%Y').date()
                        if event_date.year != year:
                            print(f"Warning: Date {date_str} is not in calendar year {year}. Skipping.")
                            continue
                        event_map[event_date.strftime('%d/%m/%Y')] = color # Single dates override ranges
                        
                    elif len(date_str.split('/')) == 2:
                        # Case 3: Short Single Date ("dd/mm")
                        event_date = datetime.datetime.strptime(f"{date_str}/{year}", '%d/%m/%Y').date()
                        event_map[event_date.strftime('%d/%m/%Y')] = color # Single dates override ranges
                        
                    else:
                        # Case 4: Unrecognized format
                        raise ValueError("Unrecognized date format.")

                except (ValueError, TypeError) as e:
                    print(f"Warning: Invalid date format '{date_str}' for '{occasion_name}'. Should be 'dd/mm/yyyy', 'dd/mm', or 'dd/mm-dd/mm'. Skipping.")
                    
    return event_map, legend_data


def draw_main_title(ws, year, styles):
    """Draws the main calendar title across the top."""
    title_cell = ws.cell(row=2, column=2) # Start at B2
    title_cell.value = f"Mosaic AI Predictive Release Calendar : {year}"
    title_cell.font = styles['main_title_font']
    title_cell.fill = styles['heading_fill']
    title_cell.alignment = styles['center_align']
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=25) 

def draw_single_month(ws, year, month, start_row, start_col, event_map, styles):
    """Draws one complete month grid at the specified location."""
    
    # 1. Month Title
    month_name = datetime.date(year, month, 1).strftime('%B').upper()
    title_cell = ws.cell(row=start_row, column=start_col)
    title_cell.value = month_name
    title_cell.font = styles['month_title_font']
    title_cell.fill = styles['heading_fill']
    title_cell.alignment = styles['center_align']
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 6)
    
    # 2. Day Headers (S, M, T, W, T, F, S)
    day_row = start_row + 1
    for i, day_abbr in enumerate(DAY_ABBREVIATIONS):
        c = ws.cell(row=day_row, column=start_col + i)
        c.value = day_abbr
        c.font = styles['day_header_font']
        c.alignment = styles['center_align']
        col_letter = get_column_letter(start_col + i)
        if ws.column_dimensions[col_letter].width is None:
             ws.column_dimensions[col_letter].width = 5
             
    # 3. Dates
    calendar.setfirstweekday(calendar.SUNDAY)
    month_weeks = calendar.monthcalendar(year, month)
    date_row_start = start_row + 2
    
    for r_idx, week in enumerate(month_weeks):
        current_row = date_row_start + r_idx
        ws.row_dimensions[current_row].height = 40
        
        for c_idx, day_num in enumerate(week):
            current_col = start_col + c_idx
            cell = ws.cell(row=current_row, column=current_col)
            cell.border = styles['date_border']

            if day_num == 0:
                cell.value = ""
                cell.fill = styles['fills']['gray']
            else:
                cell.value = day_num
                cell.alignment = styles['top_right_align']
                
                date_str = f"{day_num:02d}/{month:02d}/{year}"
                if date_str in event_map:
                    color_name = event_map[date_str]
                    cell.fill = styles['fills'][color_name]
                    
                if c_idx == 0 or c_idx == 6:
                    cell.font = styles['weekend_font']
                else:
                    cell.font = styles['date_font']

def draw_legend(ws, legend_data, styles):
    """Draws the color-coded legend to the right of the calendars."""
    
    legend_row = 6
    legend_col_color = 27 # Column AA
    legend_col_text = 28  # Column AB
    
    ws.column_dimensions[get_column_letter(legend_col_color)].width = 5
    ws.column_dimensions[get_column_letter(legend_col_text)].width = 25
    
    title_cell = ws.cell(row=legend_row, column=legend_col_text)
    title_cell.value = "Legend"
    title_cell.font = styles['legend_title_font']
    legend_row += 2
    
    for occasion, color_name in legend_data.items():
        color_cell = ws.cell(row=legend_row, column=legend_col_color)
        color_cell.fill = styles['fills'][color_name]
        color_cell.border = styles['date_border']
        
        text_cell = ws.cell(row=legend_row, column=legend_col_text)
        text_cell.value = occasion
        text_cell.font = styles['date_font']
        
        legend_row += 1

# --- Main Execution ---

def main():
    """Main function to generate the calendar."""
    print(f"Loading config from '{CONFIG_FILE}'...")
    config = load_config(CONFIG_FILE)
    year = config['calendar_year']
    
    print("Processing events and occasions...")
    event_map, legend_data = process_events(config)
    
    print("Creating new Excel workbook...")
    # --- THIS IS THE CORRECTED LINE ---
    wb = openpyxl.Workbook()
    # --- END OF CORRECTION ---
    
    ws = wb.active
    ws.title = f"Calendar {year}"
    
    styles = get_styles()
    
    draw_main_title(ws, year, styles)
    
    col_starts = [2, 10, 18] # B, J, R
    row_increment = 9
    start_row = 6
    
    print("Drawing 12-month calendar grid...")
    for month in range(1, 13):
        grid_row_index = (month - 1) // 3
        grid_col_index = (month - 1) % 3
        
        current_start_row = start_row + (grid_row_index * row_increment)
        current_start_col = col_starts[grid_col_index]
        
        draw_single_month(ws, year, month, current_start_row, current_start_col, event_map, styles)
        
    print("Drawing legend...")
    draw_legend(ws, legend_data, styles)
    
    ws.sheet_view.showGridLines = False

    try:
        print(f"Saving workbook as '{OUTPUT_FILE}'...")
        wb.save(OUTPUT_FILE)
        print(f"\nSuccess! Calendar saved to {os.path.abspath(OUTPUT_FILE)}")
    except PermissionError:
        print(f"\nError: Could not save file. '{OUTPUT_FILE}' is open or you don't have write permission.")
    except Exception as e:
        print(f"\nAn unexpected error occurred while saving: {e}")

if __name__ == "__main__":
    main()
